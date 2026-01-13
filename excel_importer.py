"""
Excel Importer module for Financial Transactions TCG
Handles importing data from Excel files
"""
import openpyxl
from typing import Tuple, List, Callable, Optional
from database import DatabaseManager


class ExcelImporter:
    def __init__(self, db_manager: DatabaseManager, progress_callback: Optional[Callable[[str], None]] = None):
        """Initialize Excel Importer with database manager"""
        self.db = db_manager
        self.imported_count = 0
        self.skipped_count = 0
        self.error_count = 0
        self.progress_callback = progress_callback
    
    def _log(self, message: str):
        """Log message to console and optionally to GUI"""
        print(message)
        if self.progress_callback:
            self.progress_callback(message)
    
    def import_file(self, file_path: str) -> Tuple[int, int, int]:
        """
        Import transactions and categories from Excel file
        Returns: (imported_count, skipped_count, error_count)
        """
        self.imported_count = 0
        self.skipped_count = 0
        self.error_count = 0
        
        try:
            wb = openpyxl.load_workbook(file_path)
            
            # Import transactions from monthly sheets (01-12)
            self._import_transactions(wb)
            
            # Import categories if "Kategorien" sheet exists
            self._import_categories(wb)
            
            wb.close()
            
        except Exception as e:
            self._log(f"Error loading Excel file: {e}")
            self.error_count += 1
        
        return (self.imported_count, self.skipped_count, self.error_count)
    
    def _import_transactions(self, workbook):
        """Import transactions from monthly sheets (01-12)"""
        # Load category mapping from Kategorien sheet
        category_map = {}
        if 'Kategorien' in workbook.sheetnames:
            cat_sheet = workbook['Kategorien']
            self._log("📂 Lade Kategorien-Mapping...")
            for row in cat_sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2 and row[0] and row[1]:
                    full_name = str(row[0]).strip()
                    short_code = str(row[1]).strip()
                    category_map[short_code] = full_name
                    self._log(f"   {short_code} -> {full_name}")
            self._log(f"✅ {len(category_map)} Kategorien geladen\n")
        
        for month_num in range(1, 13):
            sheet_name = f"{month_num:02d}"
            
            if sheet_name not in workbook.sheetnames:
                continue
            
            sheet = workbook[sheet_name]
            self._log(f"\n📄 Verarbeite Sheet: {sheet_name}")
            
            row_count = 0
            for row_idx, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
                # Skip empty rows
                if not row or all(cell is None or str(cell).strip() == '' for cell in row[:6]):
                    continue
                
                # Check if first column contains a number (ID)
                if not isinstance(row[0], (int, float)):
                    self._log(f"  Row {row_idx}: Übersprungen (keine ID in Spalte A: {row[0]})")
                    continue
                
                row_count += 1
                
                try:
                    trans_id = int(row[0])
                    date = row[1] if len(row) > 1 else None
                    description = str(row[2]) if len(row) > 2 and row[2] else ""
                    category_short = str(row[3]) if len(row) > 3 and row[3] else ""
                    
                    # Map short code to full category name
                    category = category_map.get(category_short, category_short)
                    
                    # Handle income - can be None, empty string, or number
                    income = 0.0
                    if len(row) > 4 and row[4] is not None and str(row[4]).strip():
                        try:
                            income = float(row[4])
                        except (ValueError, TypeError):
                            income = 0.0
                    
                    # Handle expense - can be None, empty string, or number
                    expense = 0.0
                    if len(row) > 5 and row[5] is not None and str(row[5]).strip():
                        try:
                            expense = float(row[5])
                        except (ValueError, TypeError):
                            expense = 0.0
                    
                    # Debug: Check if transaction exists
                    exists = self.db.transaction_exists(trans_id)
                    
                    if exists:
                        self.skipped_count += 1
                        self._log(f"  ⏭️  Row {row_idx}: ID={trans_id} übersprungen (bereits vorhanden)")
                    else:
                        # Try to insert
                        if self.db.insert_transaction(trans_id, date, description, 
                                                     category, income, expense):
                            self.imported_count += 1
                            self._log(f"  ✅ Row {row_idx}: ID={trans_id} importiert | {description[:30]} | E:{income} A:{expense}")
                        else:
                            self.skipped_count += 1
                            self._log(f"  ❌ Row {row_idx}: ID={trans_id} konnte nicht importiert werden")
                        
                except Exception as e:
                    self._log(f"  ❌ Row {row_idx}: Fehler - {str(e)}")
                    self._log(f"     Daten: {row[:6]}")
                    self.error_count += 1
            
            self._log(f"  📊 Sheet {sheet_name}: {row_count} Zeilen verarbeitet")
    
    def _import_categories(self, workbook):
        """Import categories from 'Kategorien' sheet"""
        if "Kategorien" not in workbook.sheetnames:
            return
        
        sheet = workbook["Kategorien"]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip rows without valid data in first two columns
            if not (row[0] and row[1]):
                continue
            
            try:
                category_id = str(row[0])
                label = str(row[1])
                self.db.insert_category(category_id, label)
            except Exception as e:
                self._log(f"Error importing category: {e}")
