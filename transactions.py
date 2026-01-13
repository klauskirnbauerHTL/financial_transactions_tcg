import sqlite3
import openpyxl

# SQLite-Datenbank erstellen oder verbinden
db_name = "transactions.db"
conn = sqlite3.connect(db_name)
cursor = conn.cursor()

# Tabelle transactions erstellen, falls sie noch nicht existiert
cursor.execute('''
CREATE TABLE IF NOT EXISTS transactions (
    id NUMBER PRIMARY KEY,
    date DATE,
    description TEXT,
    category TEXT,
    income REAL,
    expense REAL,
    FOREIGN KEY (category) REFERENCES categories(categoryid)
)
''')

# Tabelle categories erstellen, falls sie noch nicht existiert
cursor.execute('''
CREATE TABLE IF NOT EXISTS categories (
    categoryid TEXT PRIMARY KEY,
    label TEXT
)
''')
conn.commit()

# Benutzer nach dem Pfad der Excel-Datei fragen
excel_file = input("Bitte den Pfad zur Excel-Datei angeben: ")

# Excel-Datei laden
wb = openpyxl.load_workbook(excel_file)

# Über die Sheets "01" bis "12" iterieren
for sheet_name in [f"{i:02}" for i in range(1, 13)]:
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=7, values_only=True):  # Beginnt bei Zeile 7
            if isinstance(row[0], (int, float)):
                cursor.execute('SELECT COUNT(*) FROM transactions WHERE id = ?', (row[0],))
                if cursor.fetchone()[0] == 0:  # Prüfen, ob die ID noch nicht existiert
                    cursor.execute('''
                    INSERT INTO transactions (id, date, description, category, income, expense)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ''', row[:6])  # Nimmt die Spalten A-F
                    conn.commit()
            # Nur Zeilen mit einer ID (erste Spalte) berücksichtigen

conn.commit()

# Worksheet "Kategorien" laden und in die Tabelle categories einfügen
if "Kategorien" in wb.sheetnames:
    sheet = wb["Kategorien"]
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Beginnt bei Zeile 2
        if row[0] and row[1]:  # Nur Zeilen mit gültigen Werten in den ersten beiden Spalten
            cursor.execute('''
            INSERT OR IGNORE INTO categories (categoryid, label)
            VALUES (?, ?)
            ''', row[:2])  # Nimmt die Spalten A und B
            conn.commit()
# Verbindung schließen
conn.close()

print(f"Die Daten aus allen relevanten Sheets von {excel_file} wurden erfolgreich in die SQLite-Datenbank {db_name} importiert.")